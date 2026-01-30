from typing import List, Dict, Any  # Tipos para claridad
from openpyxl import Workbook  # Para crear Excel
from openpyxl import load_workbook  # Para leer Excel (IMPORTACIÓN)
from reportlab.lib.pagesizes import letter  # Tamaño carta para PDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # Elementos para PDF
from reportlab.lib import colors  # Colores para estilos de tabla
from reportlab.lib.styles import getSampleStyleSheet  # Estilos predefinidos de texto para PDF

CAMPOS = ["sku", "nombre", "categoria", "precio", "stock", "activo", "creado_en"]  # Columnas estándar para exportación

def exportar_excel(registros: List[Dict[str, Any]], ruta: str) -> None:  # Función para exportar a Excel
    wb = Workbook()  # Creamos un nuevo libro de Excel
    ws = wb.active  # Seleccionamos la hoja activa
    ws.title = "Productos"  # Renombramos la hoja

    ws.append([c.upper() for c in CAMPOS])  # Escribimos encabezados en la primera fila

    for r in registros:  # Recorremos los registros a exportar
        ws.append([r.get(c, "") for c in CAMPOS])  # Escribimos cada registro en una fila (mismo orden de CAMPOS)

    for col in ws.columns:  # Recorremos columnas para ajustar ancho
        max_len = 10  # Largo mínimo
        for cell in col:  # Recorremos celdas de esa columna
            val = str(cell.value) if cell.value is not None else ""  # Convertimos valor a texto
            max_len = max(max_len, len(val))  # Actualizamos máximo
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)  # Ajustamos ancho (máx 40)

    wb.save(ruta)  # Guardamos el archivo Excel en la ruta

def exportar_pdf(registros: List[Dict[str, Any]], ruta: str, titulo: str = "Reporte de Productos") -> None:  # Función para exportar a PDF
    doc = SimpleDocTemplate(ruta, pagesize=letter)  # Creamos el documento PDF tamaño carta
    styles = getSampleStyleSheet()  # Obtenemos estilos predefinidos
    story = []  # Lista de “componentes” que se agregan al PDF

    story.append(Paragraph(titulo, styles["Title"]))  # Agregamos un título
    story.append(Spacer(1, 12))  # Espacio vertical

    data = [[c.upper() for c in CAMPOS]]  # Primera fila (encabezados)
    for r in registros:  # Recorremos registros
        data.append([str(r.get(c, "")) for c in CAMPOS])  # Agregamos cada fila como texto (para evitar problemas de tipos)

    tabla = Table(data, repeatRows=1)  # Creamos tabla y repetimos encabezado en cada página
    tabla.setStyle(TableStyle([  # Definimos estilo de tabla
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),  # Fondo gris claro para encabezado
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),  # Rejilla de la tabla
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),  # Fuente negrita para encabezado
        ("FONTSIZE", (0, 0), (-1, -1), 8),  # Tamaño letra
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),  # Alineación vertical
    ]))  # Fin del estilo

    story.append(tabla)  # Agregamos tabla al PDF
    doc.build(story)  # Construimos el PDF final

def importar_excel(ruta: str) -> List[Dict[str, Any]]:  # Función para importar registros desde Excel
    wb = load_workbook(ruta)  # Abrimos el Excel
    ws = wb.active  # Tomamos la primera hoja (activa)

    headers = []  # Lista donde guardaremos los nombres de columnas del Excel
    for cell in ws[1]:  # Recorremos la primera fila (encabezados)
        headers.append(str(cell.value).strip().lower() if cell.value else "")  # Normalizamos encabezado a minúscula

    registros = []  # Lista de registros (cada uno será un dict)
    for row in ws.iter_rows(min_row=2, values_only=True):  # Recorremos filas desde la segunda (datos)
        if all(v is None or str(v).strip() == "" for v in row):  # Si la fila está completamente vacía
            continue  # La saltamos

        item = {}  # Diccionario para la fila actual
        for h, v in zip(headers, row):  # Asociamos encabezado con valor
            if not h:  # Si el encabezado está vacío
                continue  # Saltamos esa columna
            item[h] = v  # Guardamos en dict (clave=encabezado, valor=celda)

        registros.append(item)  # Agregamos la fila como dict a la lista

    return registros  # Retornamos la lista de dicts importados
